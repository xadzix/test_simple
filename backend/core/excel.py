import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import xlrd
from openpyxl import load_workbook


def _clean_cell(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _open_sheet(path, sheet_name=None):
    extension = Path(path).suffix.lower()
    if extension == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        selected_name = (
            sheet_name
            if sheet_name in workbook.sheetnames
            else workbook.sheetnames[0]
        )
        return (
            workbook,
            workbook[selected_name],
            workbook.sheetnames,
            selected_name,
            "xlsx",
        )
    if extension == ".xls":
        workbook = xlrd.open_workbook(path, on_demand=True)
        names = workbook.sheet_names()
        selected_name = sheet_name if sheet_name in names else names[0]
        return (
            workbook,
            workbook.sheet_by_name(selected_name),
            names,
            selected_name,
            "xls",
        )
    raise ValueError("Неподдерживаемый формат файла")


def _rows(sheet, kind):
    if kind == "xlsx":
        for row in sheet.iter_rows(values_only=True):
            yield [_clean_cell(value) for value in row]
    else:
        for row_index in range(sheet.nrows):
            yield [_clean_cell(value) for value in sheet.row_values(row_index)]


def read_preview(path, sheet_name=None, header_row=1, limit=15):
    workbook, sheet, names, selected_name, kind = _open_sheet(path, sheet_name)
    try:
        all_rows = _rows(sheet, kind)
        header = []
        preview_rows = []
        for number, row in enumerate(all_rows, start=1):
            if number == header_row:
                header = row
            elif number > header_row and len(preview_rows) < limit:
                preview_rows.append(row)
            elif len(preview_rows) >= limit:
                break

        width = max(
            [len(header), *(len(row) for row in preview_rows)], default=0
        )
        columns = []
        for index in range(width):
            label = (
                str(header[index]).strip()
                if index < len(header) and header[index] != ""
                else f"Колонка {index + 1}"
            )
            columns.append({"index": index, "label": label})
        normalized_rows = [
            [row[index] if index < len(row) else "" for index in range(width)]
            for row in preview_rows
        ]
        return {
            "sheet_names": names,
            "sheet_name": selected_name,
            "header_row": header_row,
            "columns": columns,
            "rows": normalized_rows,
        }
    finally:
        if kind == "xlsx":
            workbook.close()
        else:
            workbook.release_resources()


def iter_data_rows(path, sheet_name=None, header_row=1):
    workbook, sheet, _, _, kind = _open_sheet(path, sheet_name)
    try:
        for number, row in enumerate(_rows(sheet, kind), start=1):
            if number <= header_row or not any(
                value not in (None, "") for value in row
            ):
                continue
            yield number, row
    finally:
        if kind == "xlsx":
            workbook.close()
        else:
            workbook.release_resources()


def value_at(row, mapping, key, default=""):
    index = mapping.get(key)
    if index in (None, ""):
        return default
    try:
        value = row[int(index)]
    except (ValueError, TypeError, IndexError):
        return default
    return str(value).strip() if value is not None else default


def decimal_at(row, mapping, key):
    value = value_at(row, mapping, key)
    if value == "":
        return None
    normalized = re.sub(
        r"[^0-9,\.\-]", "", value.replace("\u00a0", "").replace(" ", "")
    )
    if normalized.count(",") == 1 and normalized.count(".") == 0:
        normalized = normalized.replace(",", ".")
    elif normalized.count(",") and normalized.count("."):
        normalized = normalized.replace(",", "")
    try:
        return Decimal(normalized)
    except (InvalidOperation, ValueError):
        return None
